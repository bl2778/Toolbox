import os
import tempfile
import time
import uuid
import threading
import io
import csv
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response, send_file
from werkzeug.utils import secure_filename
import google.generativeai as genai
from openai import OpenAI
from dotenv import load_dotenv
import json
from ppt_parser import extract_ppt_for_zd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here')

# Configuration
ACCESS_PASSWORD = "BAIN2025"
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'mp3', 'wav', 'm4a', 'mp4', 'mpeg', 'mpga', 'webm'}
ALLOWED_PPT_EXTENSIONS = {'pptx'}
MAX_CONTENT_LENGTH = 300 * 1024 * 1024  # 300MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Configure Gemini API
genai.configure(api_key=os.getenv('GEMINI_API_KEY'))

# Configure OpenAI API
openai_client = OpenAI(
    api_key=os.getenv('OPENAI_API_KEY'),
    base_url=os.getenv('OPENAI_BASE_URL', 'https://chat01.ai')
)

# Store transcription status in memory (in production, use Redis or database)
transcription_status = {}
# Store summary status in memory
summary_status = {}
# Store ZD analysis jobs in memory
zd_jobs = {}
zd_results = {}

# Status constants
STATUS_FILE_READING = 'file_reading'
STATUS_FILE_UPLOADED = 'file_uploaded'
STATUS_API_UPLOADING = 'api_uploading'
STATUS_API_UPLOADED = 'api_uploaded'
STATUS_PROCESSING = 'processing'
STATUS_TRANSCRIBING = 'transcribing'
STATUS_COMPLETED = 'completed'
STATUS_FAILED = 'failed'
STATUS_TIMEOUT = 'timeout'

# Summary status constants
SUMMARY_STATUS_STARTING = 'starting'
SUMMARY_STATUS_PROCESSING = 'processing'
SUMMARY_STATUS_COMPLETED = 'completed'
SUMMARY_STATUS_FAILED = 'failed'

# ZD status constants
ZD_STATUS_PARSING = 'parsing'
ZD_STATUS_CHUNKING = 'chunking'
ZD_STATUS_PROMPTING = 'prompting'
ZD_STATUS_THINKING = 'thinking'
ZD_STATUS_MERGING = 'merging'
ZD_STATUS_DONE = 'done'
ZD_STATUS_ERROR = 'error'

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_ppt_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_PPT_EXTENSIONS

def is_authenticated():
    return session.get('authenticated', False)

@app.route('/')
def index():
    if is_authenticated():
        return redirect(url_for('dashboard'))
    return render_template('auth.html')

@app.route('/auth', methods=['POST'])
def authenticate():
    password = request.form.get('password')
    if password == ACCESS_PASSWORD:
        session['authenticated'] = True
        return redirect(url_for('dashboard'))
    return render_template('auth.html', error='Invalid password')

@app.route('/dashboard')
def dashboard():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('dashboard.html')

@app.route('/transcription')
def transcription():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('transcription.html')

@app.route('/zd-tool')
def zd_tool():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('zd_tool.html')

def process_transcription_async(task_id, temp_file_path, custom_prompt, file_size, filename, model_name="gemini-2.5-pro"):
    """Async function to handle transcription processing"""
    try:
        # Update status - API uploading
        transcription_status[task_id].update({
            'status': STATUS_API_UPLOADING,
            'progress': 30,
            'message': '正在上传到Gemini API...',
            'last_update': time.time()
        })

        # Upload to Gemini
        model = genai.GenerativeModel(model_name)
        audio_file = genai.upload_file(temp_file_path)

        # Update status - API uploaded
        transcription_status[task_id].update({
            'status': STATUS_API_UPLOADED,
            'progress': 50,
            'message': 'API上传成功，等待处理...',
            'last_update': time.time()
        })

        # Wait for file to be processed with timeout
        processing_start_time = time.time()
        timeout_seconds = 300  # 5 minutes timeout

        while audio_file.state.name == "PROCESSING":
            elapsed_time = time.time() - processing_start_time

            # Check for timeout
            if elapsed_time > timeout_seconds:
                transcription_status[task_id].update({
                    'status': STATUS_TIMEOUT,
                    'message': f'处理超时（超过{timeout_seconds//60}分钟），请重新尝试',
                    'last_update': time.time()
                })
                return

            time.sleep(2)
            audio_file = genai.get_file(audio_file.name)

            # Update status with timer
            minutes = int(elapsed_time // 60)
            seconds = int(elapsed_time % 60)
            transcription_status[task_id].update({
                'status': STATUS_PROCESSING,
                'progress': 60,
                'message': f'AI正在分析音频文件... 已耗时: {minutes:02d}:{seconds:02d}',
                'elapsed_time': elapsed_time,
                'last_update': time.time()
            })

        if audio_file.state.name == "FAILED":
            transcription_status[task_id].update({
                'status': STATUS_FAILED,
                'message': 'API处理失败，请重试',
                'last_update': time.time()
            })
            return

        # Update status - transcribing with initial timer
        transcribing_start_time = time.time()
        total_elapsed = transcribing_start_time - transcription_status[task_id]['start_time']
        minutes = int(total_elapsed // 60)
        seconds = int(total_elapsed % 60)

        transcription_status[task_id].update({
            'status': STATUS_TRANSCRIBING,
            'progress': 80,
            'message': f'正在生成转写文本... 已耗时: {minutes:02d}:{seconds:02d}',
            'last_update': time.time()
        })

        # Use custom prompt if provided, otherwise use default
        default_prompt = "You are a Bain & Company consultant, just had an interview with your client, pls transcribe with timestamp"
        prompt = custom_prompt if custom_prompt.strip() else default_prompt

        # Generate transcription with periodic status updates
        def update_transcribing_status():
            while transcription_status[task_id]['status'] == STATUS_TRANSCRIBING:
                current_time = time.time()
                total_elapsed = current_time - transcription_status[task_id]['start_time']
                minutes = int(total_elapsed // 60)
                seconds = int(total_elapsed % 60)

                transcription_status[task_id].update({
                    'message': f'正在生成转写文本... 已耗时: {minutes:02d}:{seconds:02d}',
                    'last_update': current_time
                })
                time.sleep(2)  # Update every 2 seconds

        # Start timer thread for transcribing phase
        timer_thread = threading.Thread(target=update_transcribing_status)
        timer_thread.daemon = True
        timer_thread.start()

        # Generate transcription
        response = model.generate_content([audio_file, prompt])

        # Update status - completed
        total_time = time.time() - transcription_status[task_id]['start_time']
        transcription_status[task_id].update({
            'status': STATUS_COMPLETED,
            'progress': 100,
            'message': f'转写完成！总耗时: {int(total_time//60):02d}:{int(total_time%60):02d}',
            'transcription': response.text,
            'prompt_used': prompt,
            'completion_time': time.time(),
            'total_time': total_time,
            'last_update': time.time()
        })

    except Exception as e:
        # Update status - failed
        transcription_status[task_id].update({
            'status': STATUS_FAILED,
            'progress': 0,
            'message': f'处理失败: {str(e)}',
            'last_update': time.time()
        })
    finally:
        # Clean up temporary file with retry mechanism
        if temp_file_path and os.path.exists(temp_file_path):
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    os.unlink(temp_file_path)
                    break
                except PermissionError:
                    if attempt < max_retries - 1:
                        time.sleep(1)
                    else:
                        print(f"Warning: Could not delete temporary file {temp_file_path}")

@app.route('/api/transcribe', methods=['POST'])
def transcribe_audio():
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if 'audio_file' not in request.files:
        return jsonify({'error': 'No audio file provided'}), 400

    file = request.files['audio_file']
    custom_prompt = request.form.get('prompt', '')
    selected_model = request.form.get('model', 'gemini-2.5-pro')

    # Validate model selection
    valid_models = ['gemini-2.5-pro', 'gemini-2.5-flash']
    if selected_model not in valid_models:
        return jsonify({'error': f'Invalid model. Must be one of: {", ".join(valid_models)}'}), 400

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if file and allowed_file(file.filename):
        # Generate unique task ID
        task_id = str(uuid.uuid4())

        try:
            # Initialize status tracking
            transcription_status[task_id] = {
                'status': STATUS_FILE_READING,
                'progress': 5,
                'message': '文件读取中...',
                'filename': file.filename,
                'start_time': time.time(),
                'last_update': time.time()
            }

            # Create a temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1])
            temp_file_path = temp_file.name
            temp_file.close()

            # Get file size before saving
            file.seek(0, 2)
            file_size = file.tell()
            file.seek(0)

            # Save uploaded file
            file.save(temp_file_path)

            # Update status - file uploaded
            transcription_status[task_id].update({
                'status': STATUS_FILE_UPLOADED,
                'progress': 20,
                'message': '文件上传成功，准备上传到API...',
                'file_size': file_size,
                'last_update': time.time()
            })

            # Start async processing
            thread = threading.Thread(
                target=process_transcription_async,
                args=(task_id, temp_file_path, custom_prompt, file_size, file.filename, selected_model)
            )
            thread.daemon = True
            thread.start()

            # Return immediately with task ID
            return jsonify({
                'success': False,  # Processing in background
                'task_id': task_id,
                'message': 'Processing started',
                'file_info': {
                    'name': file.filename,
                    'size': file_size,
                    'processed': False
                }
            })

        except Exception as e:
            # Update status - failed
            if task_id in transcription_status:
                transcription_status[task_id].update({
                    'status': STATUS_FAILED,
                    'progress': 0,
                    'message': f'处理失败: {str(e)}',
                    'last_update': time.time()
                })
            return jsonify({'error': f'Transcription failed: {str(e)}', 'task_id': task_id}), 500

    return jsonify({'error': 'Invalid file format'}), 400

@app.route('/api/status/<task_id>')
def get_transcription_status(task_id):
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    status = transcription_status.get(task_id, {'status': 'not_found'})
    return jsonify(status)

@app.route('/api/edit/<task_id>', methods=['POST'])
def edit_transcription(task_id):
    """Edit the transcription result"""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if task_id not in transcription_status:
        return jsonify({'error': 'Task not found'}), 404

    # Check if transcription is completed
    if transcription_status[task_id]['status'] != STATUS_COMPLETED:
        return jsonify({'error': 'Transcription not completed'}), 400

    try:
        data = request.get_json()
        if not data or 'edited_transcription' not in data:
            return jsonify({'error': 'Missing edited_transcription field'}), 400

        edited_text = data['edited_transcription'].strip()
        if not edited_text:
            return jsonify({'error': 'Edited transcription cannot be empty'}), 400

        # Update the transcription
        transcription_status[task_id]['transcription'] = edited_text
        transcription_status[task_id]['edited'] = True
        transcription_status[task_id]['edit_time'] = time.time()
        transcription_status[task_id]['last_update'] = time.time()

        return jsonify({
            'success': True,
            'message': 'Transcription updated successfully',
            'transcription': edited_text
        })

    except Exception as e:
        return jsonify({'error': f'Edit failed: {str(e)}'}), 500

def process_summary_async(summary_id, transcript_text, custom_prompt, model_name):
    """Async function to handle summary generation"""
    try:
        # Update status - processing
        summary_status[summary_id].update({
            'status': SUMMARY_STATUS_PROCESSING,
            'message': '正在生成摘要...',
            'last_update': time.time()
        })

        # Prepare prompt - automatically append transcript to user's prompt
        full_prompt = custom_prompt.strip() + "\n\n<Here goes transcript>\n" + transcript_text

        # Generate summary with streaming
        response = openai_client.chat.completions.create(
            model=model_name,
            messages=[{
                "role": "user",
                "content": full_prompt
            }],
            stream=True
        )

        # Collect streaming response
        full_response = ""
        for chunk in response:
            if chunk.choices[0].delta.content is not None:
                content = chunk.choices[0].delta.content
                full_response += content

                # Update status with streaming content
                summary_status[summary_id].update({
                    'status': SUMMARY_STATUS_PROCESSING,
                    'message': '正在生成摘要...',
                    'partial_summary': full_response,
                    'last_update': time.time()
                })

        # Update status - completed
        total_time = time.time() - summary_status[summary_id]['start_time']
        summary_status[summary_id].update({
            'status': SUMMARY_STATUS_COMPLETED,
            'message': f'摘要生成完成！耗时: {int(total_time//60):02d}:{int(total_time%60):02d}',
            'summary': full_response,
            'total_time': total_time,
            'last_update': time.time()
        })

    except Exception as e:
        # Update status - failed
        summary_status[summary_id].update({
            'status': SUMMARY_STATUS_FAILED,
            'message': f'生成失败: {str(e)}',
            'last_update': time.time()
        })

@app.route('/api/summary', methods=['POST'])
def generate_summary():
    """Generate summary for transcript"""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        task_id = data.get('task_id')
        custom_prompt = data.get('prompt', '')
        model_name = data.get('model', 'gpt-4')

        # Validate task_id exists in transcription_status
        if not task_id or task_id not in transcription_status:
            return jsonify({'error': 'Invalid task ID'}), 400

        # Check if transcription is completed
        if transcription_status[task_id]['status'] != STATUS_COMPLETED:
            return jsonify({'error': 'Transcription not completed'}), 400

        # Get transcript (use edited version if available)
        transcript_text = transcription_status[task_id]['transcription']

        # Default prompt with McKinsey consultant style
        if not custom_prompt.strip():
            custom_prompt = """You are a McKinsey consultant, now helping client doing a project. Now you just had an interview, can you pls summarize the key takeaways using McKinsey wording style. Pls output in Email ready format with two sections a) Executive Summaries b) Details. The summary should exactly follow what was discussed in transcript, pls don't imagine."""

        # Validate model selection for new GPT-5 models
        valid_gpt_models = ['gpt-5', 'gpt-5-thinking', 'gpt-4', 'gpt-4-turbo', 'gpt-3.5-turbo']
        if model_name not in valid_gpt_models:
            model_name = 'gpt-5'  # Default to gpt-5

        # Generate unique summary ID
        summary_id = str(uuid.uuid4())

        # Initialize summary status
        summary_status[summary_id] = {
            'status': SUMMARY_STATUS_STARTING,
            'message': '开始生成摘要...',
            'task_id': task_id,
            'prompt_used': custom_prompt,
            'model_used': model_name,
            'start_time': time.time(),
            'last_update': time.time()
        }

        # Start async processing
        thread = threading.Thread(
            target=process_summary_async,
            args=(summary_id, transcript_text, custom_prompt, model_name)
        )
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'summary_id': summary_id,
            'message': 'Summary generation started'
        })

    except Exception as e:
        return jsonify({'error': f'Summary generation failed: {str(e)}'}), 500

@app.route('/api/summary/status/<summary_id>')
def get_summary_status(summary_id):
    """Get summary generation status"""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    status = summary_status.get(summary_id, {'status': 'not_found'})
    return jsonify(status)

@app.route('/api/summary/stream/<summary_id>')
def stream_summary(summary_id):
    """Stream summary generation progress"""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    def generate():
        last_content = ""
        while True:
            if summary_id not in summary_status:
                yield f"data: {json.dumps({'error': 'Summary not found'})}\n\n"
                break

            status = summary_status[summary_id]

            # Send status update
            yield f"data: {json.dumps(status)}\n\n"

            # If completed or failed, break the loop
            if status['status'] in [SUMMARY_STATUS_COMPLETED, SUMMARY_STATUS_FAILED]:
                break

            time.sleep(1)  # Update every second

    return Response(generate(), mimetype='text/plain')

# ZD Tool API endpoints

def process_zd_chunk_async(job_id, chunk, model_name, max_concurrency=5):
    """Process a single chunk with AI analysis."""
    chunk_id = chunk["chunk_id"]

    try:
        # Initialize chunk tracking in results
        if job_id not in zd_results:
            zd_results[job_id] = {}

        # Initialize chunk status with detailed tracking
        zd_results[job_id][chunk_id] = {
            "chunk_id": chunk_id,
            "status": "starting",
            "page_start": chunk["page_start"],
            "page_end": chunk["page_end"],
            "page_numbers": chunk["page_numbers"],
            "word_count": chunk["word_count"],
            "start_time": time.time(),
            "streaming_output": "",
            "ai_progress": "Initializing...",
            "result_text": "",
            "error": None
        }

        # Update chunk status
        if job_id not in zd_jobs:
            return

        zd_jobs[job_id]["chunks_sent"] += 1
        zd_jobs[job_id]["status"] = ZD_STATUS_THINKING
        zd_jobs[job_id]["last_update"] = time.time()

        # Update chunk status to sending
        zd_results[job_id][chunk_id]["status"] = "sending"
        zd_results[job_id][chunk_id]["ai_progress"] = "Sending request to AI..."

        # Prepare the prompt
        system_prompt = """You are a McKinsey-style consultant performing a Zero-Defect (ZD) and logic review of an English-language PowerPoint deck that has been exported for you as easy-to-read JSON.

For every slide you will receive:
• page_number
• tagline – headline text (highest priority)
• body_other – body copy, bullets, call-outs, charts (second priority; ignore purely alphanumeric labels such as "A." "I-1" that serve only as markers)
• speaker_notes – presenter notes (do not review)

Tasks for each slide:
1. Spelling mistakes – typos, repeated letters, wrong homophones, etc.
2. Grammar / phrasing issues – subject–verb agreement, tense, articles, punctuation, awkward wording, etc.
3. Logic inconsistencies:
   • Within the slide: contradictions between the tagline and the body_other content, or internal logical gaps.
   • Across slides: contradictions or mis-alignments between this slide's tagline and earlier/later taglines. (Reference both page numbers when you spot one.)

Output format:
Produce a three-column table where each row corresponds to one slide.

• Column 1 – Comma-separated list of spelling mistakes for that slide, or "—" if none.
• Column 2 – Comma-separated list of grammar / wording issues, or "—".
• Column 3 –
  "—" if the slide is logically sound.
  Otherwise, a short description of the inconsistency.
  For cross-slide issues, prefix with "↔ p X" where X is the other slide's page_number (e.g., "↔ p 5: tagline contradicts revenue trend").
• Do NOT correct or rewrite the original content; only list the issues.

**Return a GitHub-Flavored Markdown table with columns: page_number | Spelling mistakes | Grammar / wording issues | Logic inconsistencies.
No explanations, no code fences.**"""

        # Prepare user message with slides data
        user_message = "Please analyze the following slides:\n\n"
        slides_json = json.dumps(chunk["slides"], indent=2, ensure_ascii=False)
        user_message += slides_json

        # Update chunk status to processing
        zd_results[job_id][chunk_id]["status"] = "processing"
        zd_results[job_id][chunk_id]["ai_progress"] = "AI is analyzing content..."

        # Call OpenAI API with streaming
        response = openai_client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_message}
            ],
            temperature=0,
            top_p=1,
            stream=True
        )

        # Collect streaming response with real-time updates
        result_text = ""
        print(f"[DEBUG] Starting streaming for chunk {chunk_id}")

        for chunk_response in response:
            if chunk_response.choices[0].delta.content is not None:
                content = chunk_response.choices[0].delta.content
                result_text += content

                # Update streaming output in real-time
                zd_results[job_id][chunk_id]["streaming_output"] = result_text
                zd_results[job_id][chunk_id]["ai_progress"] = f"AI generating response... ({len(result_text)} chars)"
                zd_results[job_id][chunk_id]["last_update"] = time.time()

                # Update job timestamp as well
                zd_jobs[job_id]["last_update"] = time.time()

        # Final result
        result_text = result_text.strip()
        print(f"[DEBUG] Chunk {chunk_id} completed. Result length: {len(result_text)}")
        print(f"[DEBUG] First 500 chars: {result_text[:500]}...")

        zd_results[job_id][chunk_id]["result_text"] = result_text
        zd_results[job_id][chunk_id]["final_result_text"] = result_text  # Keep a separate copy
        zd_results[job_id][chunk_id]["ai_progress"] = "Analysis completed"

        print(f"[DEBUG] Chunk {chunk_id} stored in zd_results. Keys: {list(zd_results[job_id][chunk_id].keys())}")

        # Update chunk result status
        zd_results[job_id][chunk_id]["status"] = "completed"
        zd_results[job_id][chunk_id]["completion_time"] = time.time()
        zd_results[job_id][chunk_id]["ai_progress"] = f"Completed - processed {chunk['word_count']} words"

        # Update job status
        zd_jobs[job_id]["chunks_completed"] += 1
        zd_jobs[job_id]["last_update"] = time.time()

        # Check if all chunks are done
        if zd_jobs[job_id]["chunks_completed"] >= zd_jobs[job_id]["chunks_total"]:
            zd_jobs[job_id]["status"] = ZD_STATUS_MERGING
            # Trigger result merging
            merge_zd_results(job_id)

    except Exception as e:
        # Mark chunk as failed
        if job_id not in zd_results:
            zd_results[job_id] = {}

        if chunk_id not in zd_results[job_id]:
            # Initialize if not already done
            zd_results[job_id][chunk_id] = {
                "chunk_id": chunk_id,
                "page_start": chunk["page_start"],
                "page_end": chunk["page_end"],
                "page_numbers": chunk["page_numbers"],
                "word_count": chunk["word_count"],
                "start_time": time.time(),
                "streaming_output": "",
                "ai_progress": "",
                "result_text": ""
            }

        zd_results[job_id][chunk_id]["status"] = "failed"
        zd_results[job_id][chunk_id]["error"] = str(e)
        zd_results[job_id][chunk_id]["completion_time"] = time.time()
        zd_results[job_id][chunk_id]["ai_progress"] = f"Failed: {str(e)}"

        zd_jobs[job_id]["chunks_failed"] += 1
        zd_jobs[job_id]["last_update"] = time.time()

def merge_zd_results(job_id):
    """Merge results from all chunks."""
    try:
        if job_id not in zd_results or job_id not in zd_jobs:
            print(f"[DEBUG] Job {job_id} not found in results or jobs")
            return

        print(f"[DEBUG] Starting merge for job {job_id}")
        print(f"[DEBUG] Available chunks: {list(zd_results[job_id].keys())}")

        # Parse results from all chunks
        all_rows = []
        failed_chunks = []
        chunk_raw_results = {}  # Keep raw results for debugging

        for chunk_id, chunk_result in zd_results[job_id].items():
            print(f"[DEBUG] Processing chunk {chunk_id}, status: {chunk_result.get('status', 'unknown')}")

            # Store raw result for preservation
            chunk_raw_results[chunk_id] = {
                "status": chunk_result.get("status"),
                "result_text": chunk_result.get("result_text", ""),
                "final_result_text": chunk_result.get("final_result_text", ""),
                "streaming_output": chunk_result.get("streaming_output", ""),
                "page_start": chunk_result.get("page_start"),
                "page_end": chunk_result.get("page_end"),
                "word_count": chunk_result.get("word_count"),
                "error": chunk_result.get("error")
            }

            if chunk_result["status"] == "failed":
                failed_chunks.append(chunk_result)
                print(f"[DEBUG] Chunk {chunk_id} failed: {chunk_result.get('error', 'No error message')}")
                continue

            # Parse markdown table from result_text (try both sources)
            result_text = chunk_result.get("result_text", "")
            final_result_text = chunk_result.get("final_result_text", "")

            # Use final_result_text if available, otherwise fall back to result_text
            text_to_parse = final_result_text if final_result_text else result_text

            print(f"[DEBUG] Parsing chunk {chunk_id}:")
            print(f"[DEBUG]   result_text length: {len(result_text)}")
            print(f"[DEBUG]   final_result_text length: {len(final_result_text)}")
            print(f"[DEBUG]   Using text length: {len(text_to_parse)}")

            if text_to_parse:
                print(f"[DEBUG] First 200 chars: {text_to_parse[:200]}...")
                rows = parse_markdown_table(text_to_parse)
                print(f"[DEBUG] Parsed {len(rows)} rows from chunk {chunk_id}")
                all_rows.extend(rows)
            else:
                print(f"[DEBUG] No text to parse for chunk {chunk_id}")

            # Also check raw_chunk_results if available for this job
            raw_chunks = zd_jobs[job_id].get("raw_chunk_results", {})
            if chunk_id in raw_chunks and not text_to_parse:
                raw_text = raw_chunks[chunk_id].get("final_result_text", "")
                if raw_text:
                    print(f"[DEBUG] Found raw text for {chunk_id}, length: {len(raw_text)}")
                    rows = parse_markdown_table(raw_text)
                    print(f"[DEBUG] Parsed {len(rows)} rows from raw data")
                    all_rows.extend(rows)

        # Remove duplicates and sort by page number
        unique_rows = {}
        for row in all_rows:
            page_num = row.get("page_number")
            if page_num:
                if page_num not in unique_rows:
                    unique_rows[page_num] = row
                else:
                    # Merge overlapping results
                    unique_rows[page_num] = merge_row_results(unique_rows[page_num], row)

        # Sort by page number
        final_results = [unique_rows[page] for page in sorted(unique_rows.keys())]

        print(f"[DEBUG] Merged {len(final_results)} final results")
        print(f"[DEBUG] Failed chunks: {len(failed_chunks)}")

        # Update job status - PRESERVE raw chunk data
        zd_jobs[job_id]["status"] = ZD_STATUS_DONE
        zd_jobs[job_id]["final_results"] = final_results
        zd_jobs[job_id]["failed_chunks"] = failed_chunks
        zd_jobs[job_id]["completion_time"] = time.time()
        zd_jobs[job_id]["last_update"] = time.time()

        # IMPORTANT: Preserve raw chunk results for debugging and user reference
        zd_jobs[job_id]["raw_chunk_results"] = chunk_raw_results

        print(f"[DEBUG] Job {job_id} marked as DONE with {len(final_results)} results")

    except Exception as e:
        print(f"[DEBUG] Error in merge_zd_results: {str(e)}")
        zd_jobs[job_id]["status"] = ZD_STATUS_ERROR
        zd_jobs[job_id]["error"] = str(e)
        zd_jobs[job_id]["last_update"] = time.time()

def parse_markdown_table(text):
    """Parse markdown table into list of dictionaries."""
    rows = []
    lines = text.strip().split('\n')

    print(f"[DEBUG] Parsing markdown table with {len(lines)} lines")
    print(f"[DEBUG] First 10 lines: {lines[:10]}")

    # Clean up the text - remove <think> tags and other AI artifacts
    cleaned_lines = []
    in_think_block = False

    for line in lines:
        line = line.strip()

        # Skip <think> blocks
        if line.startswith('<think>'):
            in_think_block = True
            continue
        elif line.startswith('</think>'):
            in_think_block = False
            continue
        elif in_think_block:
            continue

        # Skip common AI response headers
        if line.startswith('## Answer:') or line.startswith('# Answer:') or line.startswith('Answer:'):
            continue

        if line:  # Only add non-empty lines
            cleaned_lines.append(line)

    print(f"[DEBUG] Cleaned to {len(cleaned_lines)} lines")

    # Find table start - look for header with page_number or spelling
    header_found = False
    header_index = -1

    for i, line in enumerate(cleaned_lines):
        if '|' in line and ('page_number' in line.lower() or 'spelling' in line.lower() or 'grammar' in line.lower()):
            header_found = True
            header_index = i
            print(f"[DEBUG] Found table header at line {i}: {line}")
            break

    if not header_found:
        print(f"[DEBUG] No table header found. Cleaned lines: {cleaned_lines[:5]}")
        return rows

    # Find separator line (could be |---|---| or --- | --- | ---)
    separator_index = header_index + 1
    if separator_index < len(cleaned_lines):
        separator_line = cleaned_lines[separator_index]
        print(f"[DEBUG] Separator line: {separator_line}")

        # Skip separator if it contains only dashes, pipes, and spaces
        if all(c in '-| ' for c in separator_line):
            data_start = separator_index + 1
        else:
            # No separator found, data starts right after header
            data_start = separator_index

    # Process data lines
    data_lines = cleaned_lines[data_start:]
    print(f"[DEBUG] Processing {len(data_lines)} data lines starting from index {data_start}")

    for i, line in enumerate(data_lines):
        if '|' in line and line.strip() != '':
            parts = [part.strip() for part in line.split('|')]
            print(f"[DEBUG] Line {i}: {len(parts)} parts - {parts}")

            # Handle different table formats - some may have empty first/last parts due to leading/trailing |
            if parts and parts[0] == '':
                parts = parts[1:]  # Remove empty first part
            if parts and parts[-1] == '':
                parts = parts[:-1]  # Remove empty last part

            if len(parts) >= 4:  # page_number, spelling, grammar, logic
                try:
                    # Try to parse page number from first column
                    page_str = parts[0].strip()
                    page_number = None

                    # Handle various page number formats
                    if page_str.isdigit():
                        page_number = int(page_str)
                    elif page_str.replace('.', '').isdigit():
                        page_number = int(page_str.replace('.', ''))

                    if page_number:
                        row = {
                            "page_number": page_number,
                            "spelling": parts[1] if parts[1] != "—" and parts[1] != "-" else "",
                            "grammar": parts[2] if parts[2] != "—" and parts[2] != "-" else "",
                            "logic": parts[3] if len(parts) > 3 and parts[3] != "—" and parts[3] != "-" else ""
                        }
                        rows.append(row)
                        print(f"[DEBUG] Added row for page {page_number}: {row}")
                    else:
                        print(f"[DEBUG] Could not parse page number from '{page_str}'")

                except (ValueError, IndexError) as e:
                    print(f"[DEBUG] Error parsing line {i}: {e}")
                    continue
            else:
                print(f"[DEBUG] Skipping line {i}: insufficient parts ({len(parts)})")

    print(f"[DEBUG] Parsed {len(rows)} total rows")
    return rows

def merge_row_results(row1, row2):
    """Merge results from overlapping rows."""
    merged = {"page_number": row1["page_number"]}

    for field in ["spelling", "grammar", "logic"]:
        items1 = [item.strip() for item in row1.get(field, "").split(",") if item.strip()]
        items2 = [item.strip() for item in row2.get(field, "").split(",") if item.strip()]

        # Combine and deduplicate
        combined = list(set(items1 + items2))
        merged[field] = ", ".join(combined) if combined else ""

    return merged

@app.route('/api/zd/jobs', methods=['POST'])
def create_zd_job():
    """Upload PPT file and create ZD analysis job."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if 'ppt_file' not in request.files:
        return jsonify({'error': 'No PPT file provided'}), 400

    file = request.files['ppt_file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not (file and allowed_ppt_file(file.filename)):
        return jsonify({'error': 'Invalid file format. Please upload a .pptx file'}), 400

    try:
        # Generate unique job ID
        job_id = str(uuid.uuid4())

        # Save uploaded file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pptx')
        temp_file_path = temp_file.name
        temp_file.close()
        file.save(temp_file_path)

        # Initialize job status
        zd_jobs[job_id] = {
            "job_id": job_id,
            "status": ZD_STATUS_PARSING,
            "filename": file.filename,
            "temp_file_path": temp_file_path,
            "start_time": time.time(),
            "last_update": time.time(),
            "chunks_total": 0,
            "chunks_sent": 0,
            "chunks_completed": 0,
            "chunks_failed": 0
        }

        return jsonify({
            "success": True,
            "job_id": job_id,
            "message": "File uploaded successfully"
        })

    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500

@app.route('/api/zd/jobs/<job_id>/run', methods=['POST'])
def run_zd_analysis(job_id):
    """Start ZD analysis on uploaded PPT."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if job_id not in zd_jobs:
        return jsonify({'error': 'Job not found'}), 404

    try:
        data = request.get_json() or {}
        mode = data.get('mode', 'fast')  # fast or precise
        model_name = data.get('model', 'gpt-4')

        # Validate mode
        if mode not in ['fast', 'precise']:
            mode = 'fast'

        # Validate model
        valid_models = ['gpt-5', 'gpt-5-thinking', 'gpt-4.5', 'gpt-5-pro', 'gpt-4']
        if model_name not in valid_models:
            model_name = 'gpt-4'

        job = zd_jobs[job_id]
        temp_file_path = job["temp_file_path"]

        # Update status - parsing
        job["status"] = ZD_STATUS_PARSING
        job["mode"] = mode
        job["model"] = model_name
        job["last_update"] = time.time()

        # Extract and chunk PPT
        result = extract_ppt_for_zd(temp_file_path, mode)

        if not result["success"]:
            job["status"] = ZD_STATUS_ERROR
            job["error"] = result["error"]
            job["last_update"] = time.time()
            return jsonify({"error": result["error"]}), 400

        # Update job with extracted data
        job["stats"] = result["stats"]
        job["chunks"] = result["chunks"]
        job["chunks_total"] = result["total_chunks"]
        job["status"] = ZD_STATUS_CHUNKING
        job["last_update"] = time.time()

        # Start processing chunks asynchronously
        def process_all_chunks():
            try:
                job["status"] = ZD_STATUS_PROMPTING
                job["last_update"] = time.time()

                # Process chunks with threading
                threads = []
                max_concurrent = 5

                for i, chunk in enumerate(job["chunks"]):
                    if i >= max_concurrent:
                        # Wait for previous threads to complete
                        for t in threads[:max_concurrent]:
                            t.join()
                        threads = threads[max_concurrent:]

                    thread = threading.Thread(
                        target=process_zd_chunk_async,
                        args=(job_id, chunk, model_name, max_concurrent)
                    )
                    thread.daemon = True
                    thread.start()
                    threads.append(thread)

                # Wait for all remaining threads
                for thread in threads:
                    thread.join()

            except Exception as e:
                job["status"] = ZD_STATUS_ERROR
                job["error"] = str(e)
                job["last_update"] = time.time()

        # Start async processing
        processing_thread = threading.Thread(target=process_all_chunks)
        processing_thread.daemon = True
        processing_thread.start()

        return jsonify({
            "success": True,
            "message": "Analysis started",
            "stats": result["stats"],
            "total_chunks": result["total_chunks"]
        })

    except Exception as e:
        zd_jobs[job_id]["status"] = ZD_STATUS_ERROR
        zd_jobs[job_id]["error"] = str(e)
        zd_jobs[job_id]["last_update"] = time.time()
        return jsonify({"error": f"Analysis failed: {str(e)}"}), 500

@app.route('/api/zd/jobs/<job_id>')
def get_zd_job_status(job_id):
    """Get ZD job status."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    job = zd_jobs.get(job_id, {'status': 'not_found'})

    # Calculate progress percentage
    if job.get('chunks_total', 0) > 0:
        progress = (job.get('chunks_completed', 0) / job['chunks_total']) * 100
        job['progress'] = round(progress, 1)
    else:
        job['progress'] = 0

    # Include detailed chunk information
    if job_id in zd_results:
        chunk_details = {}
        for chunk_id, chunk_data in zd_results[job_id].items():
            chunk_details[chunk_id] = {
                "chunk_id": chunk_data["chunk_id"],
                "status": chunk_data["status"],
                "page_start": chunk_data["page_start"],
                "page_end": chunk_data["page_end"],
                "word_count": chunk_data.get("word_count", 0),
                "ai_progress": chunk_data.get("ai_progress", ""),
                "streaming_output": chunk_data.get("streaming_output", ""),
                "error": chunk_data.get("error"),
                "start_time": chunk_data.get("start_time"),
                "completion_time": chunk_data.get("completion_time"),
                "last_update": chunk_data.get("last_update", time.time())
            }
        job['chunk_details'] = chunk_details

    return jsonify(job)

@app.route('/api/zd/jobs/<job_id>/result')
def get_zd_results(job_id):
    """Get ZD analysis results."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if job_id not in zd_jobs:
        return jsonify({'error': 'Job not found'}), 404

    job = zd_jobs[job_id]

    if job["status"] != ZD_STATUS_DONE:
        return jsonify({'error': 'Analysis not completed'}), 400

    format_type = request.args.get('format', 'json')

    if format_type == 'csv':
        return export_zd_results_csv(job_id)
    elif format_type == 'xlsx':
        return export_zd_results_xlsx(job_id)
    else:
        response_data = {
            "job_id": job_id,
            "status": job["status"],
            "results": job.get("final_results", []),
            "stats": job.get("stats", {}),
            "failed_chunks": job.get("failed_chunks", [])
        }

        # Include raw chunk results if requested for debugging
        include_raw = request.args.get('include_raw', 'false').lower() == 'true'
        if include_raw:
            response_data["raw_chunk_results"] = job.get("raw_chunk_results", {})

        return jsonify(response_data)

@app.route('/api/zd/jobs/<job_id>/debug')
def get_zd_debug_info(job_id):
    """Get detailed debug information for ZD analysis."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if job_id not in zd_jobs:
        return jsonify({'error': 'Job not found'}), 404

    job = zd_jobs[job_id]

    debug_info = {
        "job_id": job_id,
        "job_status": job.get("status"),
        "chunks_total": job.get("chunks_total", 0),
        "chunks_completed": job.get("chunks_completed", 0),
        "chunks_failed": job.get("chunks_failed", 0),
        "raw_chunk_results": job.get("raw_chunk_results", {}),
        "zd_results_keys": list(zd_results.get(job_id, {}).keys()) if job_id in zd_results else [],
        "final_results_count": len(job.get("final_results", [])),
        "failed_chunks_count": len(job.get("failed_chunks", []))
    }

    # Include detailed chunk info from zd_results
    if job_id in zd_results:
        debug_info["detailed_chunks"] = {}
        for chunk_id, chunk_data in zd_results[job_id].items():
            debug_info["detailed_chunks"][chunk_id] = {
                "status": chunk_data.get("status"),
                "result_text_length": len(chunk_data.get("result_text", "")),
                "streaming_output_length": len(chunk_data.get("streaming_output", "")),
                "has_final_result_text": "final_result_text" in chunk_data,
                "error": chunk_data.get("error")
            }

    return jsonify(debug_info)

@app.route('/api/zd/jobs/<job_id>/test-parse')
def test_parse_chunk(job_id):
    """Test parsing of chunk results for debugging."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if job_id not in zd_jobs:
        return jsonify({'error': 'Job not found'}), 404

    job = zd_jobs[job_id]
    test_results = {}

    # Test parsing each chunk
    raw_chunks = job.get("raw_chunk_results", {})
    for chunk_id, chunk_data in raw_chunks.items():
        result_text = chunk_data.get("final_result_text", "")
        if result_text:
            parsed_rows = parse_markdown_table(result_text)
            test_results[chunk_id] = {
                "original_length": len(result_text),
                "first_200_chars": result_text[:200],
                "parsed_rows_count": len(parsed_rows),
                "parsed_rows": parsed_rows[:3] if parsed_rows else [],  # First 3 rows for preview
                "raw_lines_count": len(result_text.split('\n'))
            }

    return jsonify({
        "job_id": job_id,
        "total_chunks": len(raw_chunks),
        "test_results": test_results
    })

def export_zd_results_csv(job_id):
    """Export results as CSV."""
    job = zd_jobs[job_id]
    results = job.get("final_results", [])

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Page Number', 'Spelling Mistakes', 'Grammar/Wording Issues', 'Logic Inconsistencies'])

    # Write data
    for row in results:
        writer.writerow([
            row["page_number"],
            row["spelling"] or "—",
            row["grammar"] or "—",
            row["logic"] or "—"
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=zd_analysis_{job_id[:8]}.csv'}
    )

def export_zd_results_xlsx(job_id):
    """Export results as Excel."""
    job = zd_jobs[job_id]
    results = job.get("final_results", [])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZD Analysis Results"

    # Write header
    headers = ['Page Number', 'Spelling Mistakes', 'Grammar/Wording Issues', 'Logic Inconsistencies']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row_idx, row in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=row["page_number"])
        ws.cell(row=row_idx, column=2, value=row["spelling"] or "—")
        ws.cell(row=row_idx, column=3, value=row["grammar"] or "—")
        ws.cell(row=row_idx, column=4, value=row["logic"] or "—")

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'zd_analysis_{job_id[:8]}.xlsx'
    )

@app.route('/api/zd/jobs/<job_id>/chunks/<chunk_id>/retry', methods=['POST'])
def retry_zd_chunk(job_id, chunk_id):
    """Retry failed chunk."""
    if not is_authenticated():
        return jsonify({'error': 'Unauthorized'}), 401

    if job_id not in zd_jobs:
        return jsonify({'error': 'Job not found'}), 404

    job = zd_jobs[job_id]

    # Find the chunk
    chunk = None
    for c in job.get("chunks", []):
        if c["chunk_id"] == chunk_id:
            chunk = c
            break

    if not chunk:
        return jsonify({'error': 'Chunk not found'}), 404

    try:
        model_name = job.get("model", "gpt-4")

        # Reset chunk status
        if job_id in zd_results and chunk_id in zd_results[job_id]:
            del zd_results[job_id][chunk_id]

        job["chunks_failed"] = max(0, job["chunks_failed"] - 1)

        # Process chunk in background
        thread = threading.Thread(
            target=process_zd_chunk_async,
            args=(job_id, chunk, model_name)
        )
        thread.daemon = True
        thread.start()

        return jsonify({"success": True, "message": "Chunk retry started"})

    except Exception as e:
        return jsonify({"error": f"Retry failed: {str(e)}"}), 500

@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    return redirect(url_for('index'))

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)