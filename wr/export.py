"""Export helpers for WR results."""

from __future__ import annotations

import csv
import io
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

from .config import EXPORT_HEADERS
from .models import ChunkResultRow


def to_csv(rows: List[ChunkResultRow]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS)
    for row in rows:
        writer.writerow([row.page, row.original, row.revised])
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(rows: List[ChunkResultRow]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Wording Revision"

    sheet.append(EXPORT_HEADERS)
    for row in rows:
        sheet.append([row.page, row.original, row.revised])

    for index, column_title in enumerate(EXPORT_HEADERS, start=1):
        column = sheet.column_dimensions[get_column_letter(index)]
        column.width = max(len(column_title) + 2, 18)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_json(rows: List[ChunkResultRow]) -> List[dict]:
    return [
        {"page": row.page, "original": row.original, "revised": row.revised}
        for row in rows
    ]
